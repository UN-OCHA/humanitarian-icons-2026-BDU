#!/usr/bin/env python3
"""
populate_metadata.py

One-time script to build metadata.json for the OCHA Humanitarian Icons 2026 project.

Reads from three sources:
  1. SVG folder  -- canonical list of icon keys (filenames without .svg)
  2. Excel file  -- display names, families, font codes, PPT flags
  3. Curator JSON -- wordmark-approved icons (43 entries) and category groupings

Outputs: metadata.json at the repo root.
"""

import json
import os
import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl

# ──────────────────────────────────────────────────────────────────────
# Paths
# ──────────────────────────────────────────────────────────────────────
BASE = Path("/Users/javiercuetoocha/OCHA DMU Dropbox/Javier Cueto/Design"
            "/Humanitarian_Icons/v2/Humanitarian_Icons_2026")
REPO = BASE / "humanitarian-icons-2026"
SVG_DIR = REPO / "svg"
EXCEL_PATH = BASE / "Humanitarian_icons_2026.xlsx"
CURATOR_PATH = BASE / "curated-icons.json"
OUTPUT_PATH = REPO / "metadata.json"

TODAY = "2026-02-13"

# Family order (from the Excel, preserved exactly)
FAMILY_ORDER = [
    "Clusters",
    "Others",
    "Disasters, hazards and crises",
    "Socioeconomic and development",
    "People",
    "Activities Strategy",
    "Product type",
    "Food and non-food items",
    "Water sanitation and hygiene",
    "Camp",
    "Security and incident",
    "Physical barriers",
    "Damage",
    "General infrastructure",
    "Logistics",
    "Telecommunications and technology",
    "UX UI",
    "Health",
    "Lockdown",
]

# ──────────────────────────────────────────────────────────────────────
# Hard-coded overrides for names that cannot be fuzzy-matched reliably
# Maps SVG key -> Excel "Icon Name" (stripped)
# ──────────────────────────────────────────────────────────────────────
MANUAL_SVG_TO_EXCEL = {
    "Camp-Coordination-and-Camp-Management":
        "Camp Coordination and Camp Management (CCCM)",
    "Anticipatory-action": "Anticipatory Action",
    "Child-care-child-friendly": "Child care Child friendly",
    "Coordinated-assessement": "Coordinated assessment",  # typo in SVG filename
    "Indigenous people": "Indigenous people",  # space in SVG filename
    "Logistics_and_Telecommunications": "Logistics and Telecommunications",
    "Multi-cluster-sector": "Multicluster sector",
    "Physical-distancing": "Social-distancing",  # renamed icon
    "Sexual-and-reproductive health": "Sexual and reproductive health",
    "Shelter-Land-and-Site-Coordination":
        "Shelter, Land and Camp Coordination",
    "Top-ranking": "Top Ranking",
}

# SVGs that exist but have NO Excel row (truly new / extra icons).
# These will get metadata entries with family="Unassigned".
# We will log them.

# ──────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    """Lowercase, strip, replace hyphens/underscores/commas with spaces,
    collapse whitespace, remove parentheticals."""
    s = s.strip().lower()
    s = re.sub(r"\(.*?\)", "", s)          # remove parentheticals
    s = s.replace("-", " ").replace("_", " ").replace(",", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def svg_key_to_normalized(key: str) -> str:
    return normalize(key)


def excel_name_to_normalized(name: str) -> str:
    return normalize(name)


# ──────────────────────────────────────────────────────────────────────
# Step 1: Scan SVG folder
# ──────────────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 1: Scanning SVG folder")
print("=" * 70)

svg_keys = sorted(
    [f[:-4] for f in os.listdir(SVG_DIR) if f.endswith(".svg")]
)
print(f"  Found {len(svg_keys)} SVG files.\n")


# ──────────────────────────────────────────────────────────────────────
# Step 2: Read Excel
# ──────────────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 2: Reading Excel file")
print("=" * 70)

wb = openpyxl.load_workbook(str(EXCEL_PATH))
ws = wb.active

excel_rows = []  # list of dicts: {family, name, font_code, ppt}
for row_idx in range(2, ws.max_row + 1):
    family = ws.cell(row_idx, 1).value
    name = ws.cell(row_idx, 2).value
    font_code = ws.cell(row_idx, 3).value
    ppt = ws.cell(row_idx, 4).value

    if family is None:
        continue  # skip blank rows

    family = family.strip()
    name = name.strip() if name else ""

    excel_rows.append({
        "family": family,
        "name": name,
        "font_code": font_code,
        "ppt": ppt,
    })

print(f"  Read {len(excel_rows)} data rows from Excel.")
print(f"  Families found: {len(set(r['family'] for r in excel_rows))}\n")


# ──────────────────────────────────────────────────────────────────────
# Step 3: Read Curator JSON
# ──────────────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 3: Reading Curator JSON")
print("=" * 70)

with open(CURATOR_PATH, "r") as f:
    curator = json.load(f)

curator_categories = curator["categories"]
curator_icons = curator["icons"]  # 43 wordmark-approved icons

# Build a set of wordmark icon names and their verticalAdjustment
# Also extract SVG key from URL where possible
wordmark_by_name = {}      # normalized display name -> curator entry
wordmark_by_svgkey = {}    # SVG key extracted from URL -> curator entry

for ci in curator_icons:
    norm = normalize(ci["name"])
    wordmark_by_name[norm] = ci

    # Extract SVG key from URL: .../SVG/UN-blue/{KEY}.svg
    url = ci.get("url", "")
    match = re.search(r"/([^/]+)\.svg$", url)
    if match:
        svgkey = match.group(1)
        wordmark_by_svgkey[svgkey] = ci

print(f"  {len(curator_icons)} wordmark-approved icons found.")
print(f"  {len(curator_categories)} categories in curator JSON.\n")


# ──────────────────────────────────────────────────────────────────────
# Step 4: Match SVG keys to Excel rows
# ──────────────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 4: Matching SVG filenames to Excel rows")
print("=" * 70)

# Build lookup from normalized Excel name -> excel_row
excel_by_norm = {}
for er in excel_rows:
    norm = excel_name_to_normalized(er["name"])
    if norm in excel_by_norm:
        print(f"  WARNING: Duplicate normalized Excel name: '{norm}' "
              f"({er['name']} vs {excel_by_norm[norm]['name']})")
    excel_by_norm[norm] = er

# Also build lookup by exact name (stripped)
excel_by_exact = {er["name"]: er for er in excel_rows}

# Match each SVG key to an Excel row
svg_to_excel = {}      # svg_key -> excel_row
unmatched_svgs = []    # svg keys with no match
matched_excel = set()  # track which excel rows got matched

for svg_key in svg_keys:
    matched = False

    # 1) Manual override
    if svg_key in MANUAL_SVG_TO_EXCEL:
        excel_name = MANUAL_SVG_TO_EXCEL[svg_key]
        if excel_name in excel_by_exact:
            svg_to_excel[svg_key] = excel_by_exact[excel_name]
            matched_excel.add(excel_name)
            matched = True
        else:
            # Try normalized
            norm = normalize(excel_name)
            if norm in excel_by_norm:
                svg_to_excel[svg_key] = excel_by_norm[norm]
                matched_excel.add(excel_by_norm[norm]["name"])
                matched = True

    if not matched:
        # 2) Try exact match: SVG key == Excel name (handles hyphenated names
        #    like "Case-management", "Exit-Cancel", "Work-from-home" etc.)
        if svg_key in excel_by_exact:
            svg_to_excel[svg_key] = excel_by_exact[svg_key]
            matched_excel.add(svg_key)
            matched = True

    if not matched:
        # 3) Normalized fuzzy match
        svg_norm = svg_key_to_normalized(svg_key)
        if svg_norm in excel_by_norm:
            svg_to_excel[svg_key] = excel_by_norm[svg_norm]
            matched_excel.add(excel_by_norm[svg_norm]["name"])
            matched = True

    if not matched:
        unmatched_svgs.append(svg_key)

# Find unmatched Excel rows
unmatched_excel = []
for er in excel_rows:
    if er["name"] not in matched_excel:
        unmatched_excel.append(er)

print(f"\n  Matched: {len(svg_to_excel)} SVG files to Excel rows.")
print(f"  Unmatched SVGs (no Excel row): {len(unmatched_svgs)}")
for s in unmatched_svgs:
    print(f"    - {s}")
print(f"  Unmatched Excel rows (no SVG): {len(unmatched_excel)}")
for er in unmatched_excel:
    print(f"    - \"{er['name']}\" (family: {er['family']})")
print()


# ──────────────────────────────────────────────────────────────────────
# Step 5: Match wordmark flags from curator JSON
# ──────────────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 5: Matching wordmark flags from Curator JSON")
print("=" * 70)

wordmark_matches = 0
wordmark_details = []

def get_wordmark_info(svg_key, display_name):
    """Return (is_wordmark, vertical_adjustment) for a given icon."""
    # Try by SVG key first (most reliable)
    if svg_key in wordmark_by_svgkey:
        ci = wordmark_by_svgkey[svg_key]
        return True, ci.get("verticalAdjustment", 0)

    # Try by normalized display name
    norm = normalize(display_name)
    if norm in wordmark_by_name:
        ci = wordmark_by_name[norm]
        return True, ci.get("verticalAdjustment", 0)

    return False, 0


for svg_key in svg_keys:
    if svg_key in svg_to_excel:
        display_name = svg_to_excel[svg_key]["name"]
    else:
        display_name = svg_key

    is_wm, va = get_wordmark_info(svg_key, display_name)
    if is_wm:
        wordmark_matches += 1
        wordmark_details.append((svg_key, display_name, va))

print(f"  Wordmark icons matched: {wordmark_matches} / {len(curator_icons)}")
if wordmark_matches != len(curator_icons):
    # Check which curator icons didn't match
    matched_curator_keys = set()
    for svg_key in svg_keys:
        if svg_key in wordmark_by_svgkey:
            matched_curator_keys.add(svg_key)
    for ci in curator_icons:
        url = ci.get("url", "")
        match = re.search(r"/([^/]+)\.svg$", url)
        if match:
            svgkey = match.group(1)
            if svgkey not in matched_curator_keys:
                # check by name
                norm = normalize(ci["name"])
                found = False
                for svg_key in svg_keys:
                    if svg_key in svg_to_excel:
                        dn = svg_to_excel[svg_key]["name"]
                    else:
                        dn = svg_key
                    if normalize(dn) == norm:
                        found = True
                        break
                if not found:
                    print(f"    UNMATCHED curator icon: {ci['name']} "
                          f"(URL key: {svgkey})")

print()
if wordmark_details:
    print("  Wordmark icons with non-zero verticalAdjustment:")
    for key, name, va in wordmark_details:
        if va != 0:
            print(f"    - {key}: verticalAdjustment={va}")
print()


# ──────────────────────────────────────────────────────────────────────
# Step 6: Assign font codepoints sequentially
# ──────────────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 6: Assigning font codepoints")
print("=" * 70)

# Alphabetical order by key (case-insensitive sort)
sorted_keys = sorted(svg_keys, key=lambda k: k.lower())

codepoint_start = 0xE001
codepoint_map = {}
for i, key in enumerate(sorted_keys):
    cp = codepoint_start + i
    codepoint_map[key] = f"U+{cp:04X}"

next_codepoint = codepoint_start + len(sorted_keys)
print(f"  Assigned {len(sorted_keys)} codepoints: "
      f"U+{codepoint_start:04X} to U+{(next_codepoint - 1):04X}")
print(f"  Next available codepoint: U+{next_codepoint:04X}\n")


# ──────────────────────────────────────────────────────────────────────
# Step 7: Build metadata.json
# ──────────────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 7: Building metadata.json")
print("=" * 70)

icons_dict = OrderedDict()

for key in sorted_keys:
    if key in svg_to_excel:
        er = svg_to_excel[key]
        display_name = er["name"]
        family = er["family"]
    else:
        # Unmatched SVG -- derive display name from key
        display_name = key.replace("-", " ").replace("_", " ")
        family = "Unassigned"

    is_wm, va = get_wordmark_info(key, display_name)

    icons_dict[key] = OrderedDict([
        ("name", display_name),
        ("family", family),
        ("wordmark", is_wm),
        ("wordmark_valign", va),
        ("font_codepoint", codepoint_map[key]),
        ("date_added", TODAY),
    ])

metadata = OrderedDict([
    ("meta", OrderedDict([
        ("version", "2.0"),
        ("last_updated", TODAY),
        ("next_font_codepoint", f"U+{next_codepoint:04X}"),
    ])),
    ("families", FAMILY_ORDER),
    ("icons", icons_dict),
])

# Write the file
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(metadata, f, indent=2, ensure_ascii=False)

print(f"  Written {len(icons_dict)} icon entries to {OUTPUT_PATH}")
print(f"  File size: {os.path.getsize(OUTPUT_PATH):,} bytes\n")


# ──────────────────────────────────────────────────────────────────────
# Summary report
# ──────────────────────────────────────────────────────────────────────
print("=" * 70)
print("MATCH REPORT SUMMARY")
print("=" * 70)
print(f"  SVG files scanned:         {len(svg_keys)}")
print(f"  Excel data rows:           {len(excel_rows)}")
print(f"  Curator wordmark icons:    {len(curator_icons)}")
print()
print(f"  SVGs matched to Excel:     {len(svg_to_excel)}")
print(f"  SVGs unmatched (no Excel): {len(unmatched_svgs)}")
for s in unmatched_svgs:
    print(f"    - {s}")
print(f"  Excel rows unmatched:      {len(unmatched_excel)}")
for er in unmatched_excel:
    print(f"    - \"{er['name']}\" (family: {er['family']})")
print()
print(f"  Wordmark icons matched:    {wordmark_matches} / {len(curator_icons)}")
print(f"  Font codepoints assigned:  {len(codepoint_map)}")
print(f"  Families in output:        {len(FAMILY_ORDER)}")
print()

# Validate: every SVG has an entry
assert len(icons_dict) == len(svg_keys), \
    f"Icon count mismatch: {len(icons_dict)} vs {len(svg_keys)} SVGs"

# Check for any "Unassigned" families
unassigned = [k for k, v in icons_dict.items() if v["family"] == "Unassigned"]
if unassigned:
    print(f"  WARNING: {len(unassigned)} icons with family='Unassigned':")
    for k in unassigned:
        print(f"    - {k} (display: {icons_dict[k]['name']})")
    print()

print("Done.")
