#!/usr/bin/env python3
"""
Generate an Excel spreadsheet from the OCHA humanitarian icons metadata.

Reads metadata.json and produces a clean, shareable .xlsx file with one row
per icon, grouped by family (in the order defined in metadata.families) and
sorted alphabetically within each family.

Usage:
    python scripts/generate-excel.py

Output:
    output/Humanitarian_icons.xlsx
"""

import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent
METADATA_PATH = REPO_ROOT / "metadata.json"
OUTPUT_PATH = REPO_ROOT / "output" / "Humanitarian_icons.xlsx"


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="1F3864")
CELL_FONT = Font(name="Calibri", size=11)
THIN_BORDER_SIDE = Side(style="thin", color="BFC9D6")
HEADER_BORDER = Border(
    bottom=Side(style="medium", color="8FAADC"),
)
CELL_BORDER = Border(
    bottom=Side(style="thin", color="D6DCE4"),
)


def load_metadata() -> dict:
    """Load and return the parsed metadata.json."""
    with open(METADATA_PATH, encoding="utf-8") as f:
        return json.load(f)


def build_rows(metadata: dict) -> list[list[str]]:
    """
    Build a list of row data from the metadata, respecting family order
    and alphabetical sorting within each family.

    Each row: [Family, Icon name, Date added]
    """
    families_order = metadata["families"]
    icons = metadata["icons"]

    # Group icons by family
    family_icons: dict[str, list[dict]] = {f: [] for f in families_order}
    for _key, icon in icons.items():
        family = icon["family"]
        if family in family_icons:
            family_icons[family].append(icon)
        else:
            # Safety net for icons whose family isn't listed
            family_icons.setdefault(family, []).append(icon)

    rows: list[list[str]] = []
    for family in families_order:
        sorted_icons = sorted(family_icons[family], key=lambda i: i["name"].lower())
        for icon in sorted_icons:
            rows.append([
                family,
                icon["name"],
                icon.get("date_added", ""),
            ])

    # Append any icons from families not in the canonical order
    known_families = set(families_order)
    for family, icons_list in family_icons.items():
        if family not in known_families:
            sorted_icons = sorted(icons_list, key=lambda i: i["name"].lower())
            for icon in sorted_icons:
                rows.append([
                    family,
                    icon["name"],
                    icon.get("date_added", ""),
                ])

    return rows


def auto_fit_columns(ws, headers: list[str], rows: list[list[str]]) -> None:
    """Set each column width to fit the widest cell content, with padding."""
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            cell_len = len(str(row[col_idx - 1]))
            if cell_len > max_len:
                max_len = cell_len
        # Add padding (2 characters) for comfortable reading
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4


def create_workbook(rows: list[list[str]]) -> Workbook:
    """Create and return a styled openpyxl Workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Humanitarian Icons"

    headers = ["Family", "Icon name", "Date added"]

    # --- Header row ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(vertical="center")

    # --- Data rows ---
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center")

    # --- Column widths ---
    auto_fit_columns(ws, headers, rows)

    # Freeze header row for easier scrolling
    ws.freeze_panes = "A2"

    # Set print title (repeat header on every printed page)
    ws.print_title_rows = "1:1"

    return wb


def main() -> None:
    print(f"Reading metadata from {METADATA_PATH}")
    metadata = load_metadata()

    families = metadata["families"]
    icons = metadata["icons"]
    print(f"  Found {len(icons)} icons across {len(families)} families")

    rows = build_rows(metadata)

    wb = create_workbook(rows)

    # Ensure output directory exists
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(OUTPUT_PATH))
    print(f"  Wrote {len(rows)} data rows to {OUTPUT_PATH}")
    print("Done.")


if __name__ == "__main__":
    main()
